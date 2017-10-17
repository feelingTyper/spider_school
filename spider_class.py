# -*- coding: UTF-8 -*-

import time
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
import xlrd
import genius
import logging
import logging.handlers

# reload(sys)
# sys.setdefaultencoding('utf8')


class Spider(object):
    # Some User Agents
    # 请求头，现在没用到
    hds = [{'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6', \
            "Cookie": "PHPSESSID=sud248vrpjn3u0pfr088cjjo61; sgsa_id=cntoubang.com|1499326402758665; qs_lvt_9076=1499326402; sgsa_vt_unknown=1499335837255; qs_pv_9076=1920476421293224400%2C2281014433573530400%2C1049934143332009200%2C811280901206671500%2C260639860455639780"
        }, \
           {
               'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11', \
               "Cookie": "PHPSESSID=sud248vrpjn3u0pfr088cjjo61; sgsa_id=cntoubang.com|1499326402758665; qs_lvt_9076=1499326402; sgsa_vt_unknown=1499335837255; qs_pv_9076=1920476421293224400%2C2281014433573530400%2C1049934143332009200%2C811280901206671500%2C260639860455639780"
           }, \
           {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)', \
            "Cookie": "PHPSESSID=sud248vrpjn3u0pfr088cjjo61; sgsa_id=cntoubang.com|1499326402758665; qs_lvt_9076=1499326402; sgsa_vt_unknown=1499335837255; qs_pv_9076=1920476421293224400%2C2281014433573530400%2C1049934143332009200%2C811280901206671500%2C260639860455639780"
            },\
           {
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36', \
            "Cookie": "ASP.NET_SessionId=zwtoidiyecbexuwxwophzzmo; __utmt=1; sc_pview_shuser=; _ga=GA1.2.453946955.1507714115; _gid=GA1.2.198735875.1507714116;\
             __utma=191177727.453946955.1507714115.1507788618.1507790873.4; __utmb=191177727.2.10.1507790873; __utmc=191177727; __utmz=191177727.1507788618.3.2.utmcsr=mail.qq.com|\
             utmccn=(referral)|utmcmd=referral|utmcct=/"
           }]
    filename = "name_dict.xlsx"
    name_dict = ''

    def __init__(self):
        '''
        从拼音文件中读取所有的中国姓氏到list中
        '''
        self.name_dict = self.load_py(self.filename)

    def getlog(self, logfile, logname):
        '''
            产生日志对象，包括日志文件的名称，和日志的名称，主要用来记录当前爬行的进度
        '''
        # LOG_FILE = 'spider.log'
        handler = logging.handlers.RotatingFileHandler(logfile, maxBytes=1024 * 1024, backupCount=5)  # 实例化handler
        fmt = '%(asctime)s - %(filename)s:%(lineno)s - %(name)s - %(message)s'

        formatter = logging.Formatter(fmt)  # 实例化formatter
        handler.setFormatter(formatter)  # 为handler添加formatter

        logger = logging.getLogger(logname)  # 获取名为tst的logger
        logger.addHandler(handler)  # 为logger添加handler
        logger.setLevel(logging.DEBUG)

        return logger

    def load_py(self, filename):
        '''
            从拼音文件中读取所有的中国姓氏到list中
        '''
        xls = xlrd.open_workbook(filename)
        sheet = xls.sheets()[0]
        col = sheet.col_values(0)
        return col

    def is_chinese(self, text, namedict):
        '''
        辨别是否是中国学生，是返回1，不是返回0
        :param text:
        :param namedict:
        :return:
        '''
        text = text.lower()
        text = text.replace(' ', '')
        print('Text is %s' % text)
        if not text:
            return 0

        count = 0
        try:
            seg_list = genius.seg_text(
                text,
                use_combine=True,
                use_pinyin_segment=True,
                use_tagging=True,
                use_break=True
            )
            for pinyin in seg_list:
                print('pinyin', pinyin.text)
                if pinyin.text in namedict:
                    count += 1

            if count/len(seg_list) > 0.5:
                return 1
            else:
                return 0
        except IndexError as e:
            print('error in genius')
            print(e)
        return 0

    def parse_content(self, content):
        '''
        解析html内容的函数，没用到
        :param content:
        :return:
        '''
        if content is not str:
            content = str(content)
        soup = BeautifulSoup(content, "html.parser")
        return soup.get_text()

    def print_book_lists_excel_list(self, student_lists, student_tag_lists, save_path):
        '''
        将list中的数据保存到excel中去,按照拼音来存储

        :param student_lists:
        :param student_tag_lists:
        :param save_path:
        :return:
        '''
        if not os.path.exists(save_path):
            os.makedirs(save_path)

        wb = Workbook()
        ws = []

        for i in range(len(student_tag_lists)):
            ws.append(wb.create_sheet(title=student_tag_lists[i]))  # utf8->unicode

        for i in range(len(student_tag_lists)):
            ws[i].append(['序号', '姓名', '年级', '邮箱', '是否为中国学生'])
            count = 1
            for sl in student_lists[i]:
                ws[i].append([count, sl[1], sl[2], sl[3], sl[4]])
                count += 1

        save_path += 'total_py.xlsx'
        wb.save(save_path)

    def print_book_lists_excel(self, student_lists, student_tag, save_path):
        '''
        将list中的数据保存到excel中去
        :param student_lists:
        :param student_tag:
        :param save_path:
        :return:
        '''
        if not os.path.exists(save_path):
            os.mkdirs(save_path)

        wb = Workbook()
        ws = wb.create_sheet(title=student_tag)  # utf8->unicode
        ws.append(['序号', '姓名', '年级', '邮箱', '是否为中国学生'])
        count = 1
        for bl in student_lists:
            ws.append([count, bl[0], bl[1], bl[2], bl[3]])
            count += 1

        save_path += student_tag
        save_path += '.xlsx'
        wb.save(save_path)

    def get_count(self, filename):
        '''
        从日志文件中读取当前爬取的位置
        :param filename:
        :return:
        '''
        if not os.path.exists(filename):
            return 0

        with open(filename, 'r') as f:
            lines = f.readlines()
            if not lines:
                return 0
            last_line = lines[-1]
            words = last_line.split('-')

        count = words[-1]
        return int(count)

    def get_sheet(self, filename):
        '''
        读取xlsx文件中的数据到内存中，并返回数据和sheet名称（拼音名）
        :param filename:
        :return:
        '''
        student_lists = []
        xls = xlrd.open_workbook(filename)
        sheet = xls.sheets()[1]
        for i in range(1, sheet.nrows):
            student_lists.append(sheet.row_values(i))
        return student_lists, sheet.name

    def get_xlsx_date(self, save_path):
        '''
        读取所有的xlsx文件内容
        :param save_path:
        :return:
        '''
        if not os.path.exists(save_path):
            return 'error in compact_xlsx'

        sheet_lists = []
        sheet_names = []

        pathDir = os.listdir(save_path)

        for file in pathDir:
            file_path = save_path + file
            sheet_list, sheet_name = self.get_sheet(file_path)
            sheet_lists.append(sheet_list)
            sheet_names.append(sheet_name)

        return sheet_lists, sheet_names

    def compct_xlsx_py(self, save_path, output_path):
        '''
        按照拼音合并
        :param save_path:
        :param output_path:
        :return:
        '''
        sheet_lists, sheet_names = self.get_xlsx_date(save_path)
        self.print_book_lists_excel_list(sheet_lists, sheet_names, output_path)

    def compct_xlsx_all(self, save_path, output_path):
        '''
        所有的xlsx内容合并到一张表中
        :param save_path:
        :param output_path:
        :return:
        '''
        sheet_all = []
        sheet_lists, sheet_names = self.get_xlsx_date(save_path)
        for sheet_list in sheet_lists:
            for sheet in sheet_list:
                sheet = sheet[1:]
                sheet_all.append(sheet)
        self.print_book_lists_excel(sheet_all, 'total_all', output_path)

    def compct_xlsx_all_chinese(self, save_path, output_path):
        '''
        只要中国学生
        :param save_path:
        :param output_path:
        :return:
        '''
        sheet_all = []
        sheet_lists, sheet_names = self.get_xlsx_date(save_path)
        final_path = save_path
        for sheet_list in sheet_lists:
            for sheet in sheet_list:
                if not sheet[-1]:
                    continue
                sheet = sheet[1:]
                sheet_all.append(sheet)
        self.print_book_lists_excel(sheet_all, 'total_chinese', output_path)

